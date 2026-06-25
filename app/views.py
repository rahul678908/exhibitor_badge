from django.contrib.auth import authenticate

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.pagination import PageNumberPagination
from rest_framework.generics import UpdateAPIView, DestroyAPIView
from rest_framework import status, serializers
from django.db.models import Sum, Count
from rest_framework_simplejwt.tokens import RefreshToken
from django.db import transaction
from .pagination import RegistrationPagination
from django.core.cache import cache
import io
from rest_framework.exceptions import ValidationError
from django.shortcuts import get_object_or_404
from rest_framework.parsers import MultiPartParser, FormParser
import pandas as pd
from .tasks import process_bulk_upload
from django.db.models import F
from .tasks import process_bulk_upload, commit_bulk_upload
from .utils import verify_recaptcha
from .utils import get_exhibitor_allocation

from .models import (
    TicketType,
    BadgeAllocation,
    Exhibitor,
    User,
    Registration,
    Invitation,
    UploadBatch,
    UploadBatchRecord,
    UploadFieldMapping,
)

from rest_framework.generics import (
    ListAPIView, 
    RetrieveAPIView,
    CreateAPIView,
)

from .serializers import (
    CreateExhibitorSerializer,
    ExhibitorListSerializer,
    ExhibitorDetailSerializer,
    ExhibitorLoginSerializer, 
    TicketTypeCreateSerializer, 
    TicketTypeUpdateSerializer, 
    TicketTypeListSerializer, 
    BadgeAllocationSerializer, 
    ExhibitorCreateSerializer, 
    RegistrationListSerializer,
    LoginSerializer,
    RegistrationCreateSerializer,
    RegistrationUpdateSerializer,
    UploadBatchSerializer,
    UploadBatchListSerializer,
    UploadBatchRecordSerializer,
    RecordEditSerializer,
    BadgeAllocationListSerializer

)




 






class BulkUploadPagination(
    PageNumberPagination
):

    page_size = 100

    page_size_query_param = "page_size"

    max_page_size = 500





class ExhibitorMyAllocationsAPIView(ListAPIView):
    """
    GET /exhibitor/my-allocations/
    Returns the calling exhibitor's BadgeAllocation rows with usage info.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = BadgeAllocationListSerializer   # already defined in your codebase
 
    def get_queryset(self):
        exhibitor = self.request.user.exhibitor_profile
        return BadgeAllocation.objects.filter(
            exhibitor=exhibitor
        ).select_related("ticket_type")



class AdminExhibitorAllocationsAPIView(ListAPIView):
    """
    GET /admin/exhibitors/<exhibitor_id>/allocations/
    Returns all BadgeAllocation rows for the given exhibitor,
    augmented with per-ticket-type global pool info for the allocation panel.
    """
    permission_classes = [IsAuthenticated]
    serializer_class = BadgeAllocationListSerializer
 
    def get_queryset(self):
        return BadgeAllocation.objects.filter(
            exhibitor_id=self.kwargs["exhibitor_id"]
        ).select_related("ticket_type")



class CreateOrUpdateBadgeAllocationAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        serializer = BadgeAllocationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        exhibitor = get_object_or_404(Exhibitor, id=data["exhibitor_id"])
        ticket_type = get_object_or_404(
            TicketType.objects.select_for_update(), id=data["ticket_type_id"]
        )

        # ✅ Don't let admin hand out more than the global pool has left
        already_allocated_elsewhere = BadgeAllocation.objects.filter(
            ticket_type=ticket_type
        ).exclude(exhibitor=exhibitor).aggregate(
            total=Sum("allocated_count")
        )["total"] or 0

        if already_allocated_elsewhere + data["allocated_count"] > ticket_type.total_tickets:
            remaining = ticket_type.total_tickets - already_allocated_elsewhere
            return Response(
                {
                    "status": False,
                    "message": (
                        f"Cannot allocate {data['allocated_count']} '{ticket_type.ticket_name}' badges. "
                        f"Only {remaining} remain unallocated out of {ticket_type.total_tickets} total "
                        f"({already_allocated_elsewhere} already allocated to other exhibitors)."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        allocation, created = BadgeAllocation.objects.select_for_update().get_or_create(
            exhibitor=exhibitor,
            ticket_type=ticket_type,
            defaults={
                "allocated_count": data["allocated_count"],
                "allocated_by": request.user,
                "remarks": data.get("remarks", ""),
            }
        )

        if not created:
            used = Registration.objects.filter(
                exhibitor=exhibitor, ticket_type=ticket_type
            ).exclude(status="cancelled").count()

            if data["allocated_count"] < used:
                return Response(
                    {
                        "status": False,
                        "message": (
                            f"Cannot reduce allocation to {data['allocated_count']}. "
                            f"{used} badge(s) are already in use against it."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            allocation.allocated_count = data["allocated_count"]
            allocation.allocated_by = request.user
            allocation.remarks = data.get("remarks", allocation.remarks)
            allocation.save()

        return Response(
            {
                "status": True,
                "message": "Badge allocation saved successfully",
                "allocation": {
                    "id": allocation.id,
                    "exhibitor": exhibitor.company_name,
                    "ticket_type": ticket_type.ticket_name,
                    "allocated_count": allocation.allocated_count,
                }
            },
            status=status.HTTP_200_OK,
        )


class ExhibitorBadgeAllocationListAPIView(ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = BadgeAllocationListSerializer

    def get_queryset(self):
        return BadgeAllocation.objects.filter(
            exhibitor_id=self.kwargs["exhibitor_id"]
        ).select_related("ticket_type")


# ─────────────────────────────────────────────
# INVITATION VIEWS
# Add these to your existing views.py
# ─────────────────────────────────────────────

# POST /exhibitor/invitations/send/
# Creates Registration (status=invited) + Invitation (token) for each
class SendInvitationAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        exhibitor = get_exhibitor(request)
        entries = request.data.get("entries", [])

        if not entries:
            return Response(
                {"error": "No entries provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = []
        errors = []

        # ✅ NEW — running per-ticket-type usage for THIS exhibitor, so
        # multiple entries for the same ticket type in one request can't
        # blow past the allocation before any row is even committed.
        allocation_state = {}

        def get_state(ticket_type_obj):
            if ticket_type_obj.id not in allocation_state:
                allocated, used, _ = get_exhibitor_allocation(exhibitor, ticket_type_obj)
                allocation_state[ticket_type_obj.id] = {"allocated": allocated, "used": used}
            return allocation_state[ticket_type_obj.id]

        for idx, entry in enumerate(entries):
            first_name = entry.get("first_name", "").strip()
            last_name = entry.get("last_name", "").strip()
            email = entry.get("email", "").strip()
            ticket_type_id = entry.get("ticket_type_id")

            row_errors = []

            if not first_name:
                row_errors.append("First name is required.")

            if not last_name:
                row_errors.append("Last name is required.")

            if not email:
                row_errors.append("Email is required.")

            if not ticket_type_id:
                row_errors.append("Ticket type is required.")

            if row_errors:
                errors.append({
                    "row": idx + 1,
                    "email": email,
                    "errors": row_errors,
                })
                continue

            if Registration.objects.filter(email=email).exists():
                errors.append({
                    "row": idx + 1,
                    "email": email,
                    "errors": ["Email already registered."],
                })
                continue

            try:
                ticket = TicketType.objects.get(id=ticket_type_id)
            except TicketType.DoesNotExist:
                errors.append({
                    "row": idx + 1,
                    "email": email,
                    "errors": ["Invalid ticket type."],
                })
                continue

            # ── Check Exhibitor's Own Allocation ────────────  ✅ NEW
            state = get_state(ticket)
            exhibitor_available = state["allocated"] - state["used"]

            if state["allocated"] == 0 or exhibitor_available <= 0:
                errors.append({
                    "row": idx + 1,
                    "email": email,
                    "errors": [
                        f"You have only {max(exhibitor_available, 0)} '{ticket.ticket_name}' "
                        f"badge(s) remaining out of your allocated {state['allocated']}. "
                        "Please contact the administrator to increase your quota."
                    ],
                })
                continue

            # ── Check Global Ticket Availability ────────────
            used_count = Registration.objects.filter(
                ticket_type=ticket
            ).exclude(
                status="cancelled"
            ).count()

            available_count = (
                ticket.total_tickets - used_count
            )

            if available_count <= 0:
                errors.append({
                    "row": idx + 1,
                    "email": email,
                    "errors": [
                        f"No available badges left for '{ticket.ticket_name}'."
                    ],
                })
                continue

            try:
                with transaction.atomic():

                    reg = Registration.objects.create(
                        exhibitor=exhibitor,
                        ticket_type=ticket,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        registered_via="invitation",
                        status="invited",
                    )

                    invitation = Invitation.objects.create(
                        registration=reg,
                        status="sent",
                    )

                    state["used"] += 1  # ✅ NEW — reflect this entry for subsequent rows

                    created.append({
                        "id": reg.id,
                        "name": f"{first_name} {last_name}",
                        "email": email,
                        "invitation_link": invitation.invitation_link,
                        "token": invitation.token,
                    })

            except Exception as e:
                errors.append({
                    "row": idx + 1,
                    "email": email,
                    "errors": [str(e)],
                })

        return Response(
            {
                "created": len(created),
                "failed": len(errors),
                "results": created,
                "errors": errors,
            },
            status=status.HTTP_201_CREATED,
        )

# # POST /exhibitor/invitations/import/
# # Upload CSV/Excel to import contacts for invitation
# class InvitationImportFileView(APIView):
#     permission_classes = [IsAuthenticated]
#     parser_classes = [MultiPartParser, FormParser]

#     def post(self, request):
#         file = request.FILES.get("file")

#         if not file:
#             return Response(
#                 {"error": "No file provided."},
#                 status=status.HTTP_400_BAD_REQUEST,
#             )

#         try:
#             fname = file.name.lower()
#             if fname.endswith(".csv"):
#                 df = pd.read_csv(file)
#             elif fname.endswith((".xlsx", ".xls")):
#                 df = pd.read_excel(file, engine="openpyxl")
#             else:
#                 return Response(
#                     {"error": "Only CSV and Excel files are supported."},
#                     status=status.HTTP_400_BAD_REQUEST,
#                 )
#         except Exception as e:
#             return Response(
#                 {"error": f"Failed to read file: {str(e)}"},
#                 status=status.HTTP_400_BAD_REQUEST,
#             )

#         df = df.fillna("")

#         # Normalize column names
#         df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

#         entries = []
#         for _, row in df.iterrows():
#             entries.append({
#                 "first_name": str(row.get("first_name", row.get("firstname", ""))).strip(),
#                 "last_name": str(row.get("last_name", row.get("lastname", ""))).strip(),
#                 "email": str(row.get("email", row.get("email_address", ""))).strip(),
#                 "ticket_type_name": str(row.get("ticket_type", "")).strip(),
#             })

#         return Response({"entries": entries, "total": len(entries)})


# GET /exhibitor/invitations/
# List all invitations for the exhibitor
class InvitationListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        exhibitor = get_exhibitor(request)

        registrations = Registration.objects.filter(
            exhibitor=exhibitor,
            registered_via="invitation",
        ).select_related("ticket_type", "invitation").order_by("-created_at")

        data = []
        for reg in registrations:
            inv = getattr(reg, "invitation", None)
            data.append({
                "id": reg.id,
                "urn": reg.urn,
                "first_name": reg.first_name,
                "last_name": reg.last_name,
                "full_name": f"{reg.first_name} {reg.last_name}",
                "email": reg.email,
                "ticket_type": reg.ticket_type.ticket_name if reg.ticket_type else "",
                "status": reg.status,
                "invitation_link": inv.invitation_link if inv else None,
                "invitation_status": inv.status if inv else None,
                "sent_at": inv.sent_at if inv else None,
            })

        return Response(data)


# GET /register/<token>/   — public, no auth
# Returns registration details for the attendee to complete
class InvitationRegisterDetailView(APIView):
    permission_classes = []
    authentication_classes = []

    def get(self, request, token):
        invitation = get_object_or_404(Invitation, token=token)
        reg = invitation.registration

        if reg.status == "confirmed":
            return Response(
                {"error": "This invitation has already been completed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Mark as opened
        if invitation.status == "sent":
            from django.utils import timezone
            invitation.status = "opened"
            invitation.opened_at = timezone.now()
            invitation.save(update_fields=["status", "opened_at"])

        return Response({
            "token": token,
            "first_name": reg.first_name,
            "last_name": reg.last_name,
            "email": reg.email,
            "ticket_type": reg.ticket_type.ticket_name if reg.ticket_type else "",
            "status": reg.status,
        })


# POST /register/<token>/complete/  — public, no auth
# Attendee submits remaining details → status becomes confirmed
class InvitationRegisterCompleteView(APIView):
    permission_classes = []
    authentication_classes = []

    def post(self, request, token):

        invitation = get_object_or_404(Invitation, token=token)
        reg = invitation.registration

        if reg.status == "confirmed":
            return Response(
                {"error": "Already completed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check ticket availability
        used_count = Registration.objects.filter(
            ticket_type=reg.ticket_type
        ).exclude(
            status="cancelled"
        ).exclude(
            id=reg.id
        ).count()

        available_count = reg.ticket_type.total_tickets - used_count

        if available_count <= 0:
            return Response(
                {
                    "error": (
                        f"No available badges left for "
                        f"'{reg.ticket_type.ticket_name}'."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ↓ These were mistakenly indented inside the if block above
        reg.job_title = request.data.get("job_title", reg.job_title)
        reg.company_name = request.data.get("company_name", reg.company_name)
        reg.phone_number = request.data.get("phone_number", reg.phone_number)
        reg.country_of_residence = request.data.get("country_of_residence", reg.country_of_residence)
        reg.nationality = request.data.get("nationality", reg.nationality)
        reg.terms_accepted = request.data.get("terms_accepted", False)
        reg.status = "confirmed"
        reg.save()

        from django.utils import timezone

        invitation.status = "completed"
        invitation.completed_at = timezone.now()
        invitation.save(update_fields=["status", "completed_at"])

        return Response(
            {
                "message": "Registration completed successfully.",
                "urn": reg.urn,
                "name": f"{reg.first_name} {reg.last_name}",
                "email": reg.email,
            }
        )

class InvitationRegisterUpdateNameView(APIView):
    permission_classes = []
    authentication_classes = []

    def put(self, request, token):
        invitation = get_object_or_404(Invitation, token=token)
        reg = invitation.registration

        if reg.status == "confirmed":
            return Response(
                {"error": "This invitation has already been completed."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        first_name = request.data.get("first_name", "").strip()
        last_name  = request.data.get("last_name", "").strip()

        if not first_name:
            return Response({"error": "First name is required."}, status=status.HTTP_400_BAD_REQUEST)
        if not last_name:
            return Response({"error": "Last name is required."}, status=status.HTTP_400_BAD_REQUEST)

        reg.first_name = first_name
        reg.last_name  = last_name
        reg.save(update_fields=["first_name", "last_name"])

        return Response({
            "first_name": reg.first_name,
            "last_name":  reg.last_name,
        })








 
 
 


def get_exhibitor(request):
    return get_object_or_404(Exhibitor, user=request.user)
 
 
SYSTEM_FIELDS = [
    {"key": "first_name",           "label": "First Name",           "required": True},
    {"key": "last_name",            "label": "Last Name",            "required": True},
    {"key": "email",                "label": "Email",                "required": True},
    {"key": "job_title",            "label": "Job Title",            "required": False},
    {"key": "company_name",         "label": "Company Name",         "required": False},
    {"key": "phone_number",         "label": "Phone Number",         "required": False},
    {"key": "country_of_residence", "label": "Country Of Residence", "required": False},
    {"key": "nationality",          "label": "Nationality",          "required": False},
    {"key": "ticket_type",          "label": "Ticket Type",          "required": True},
]
 
REQUIRED_TARGET_FIELDS = {"first_name", "last_name", "email", "ticket_type"}
 
 
# ─────────────────────────────────────────────
# STEP 1 — Upload file, return columns preview
# POST /exhibitor/bulk-upload/upload/
# ─────────────────────────────────────────────
class BulkUploadFileView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
 
    def post(self, request):
        file = request.FILES.get("file")
        batch_name = request.data.get("batch_name", "").strip()
 
        if not file:
            return Response(
                {"error": "No file provided."},
                status=status.HTTP_400_BAD_REQUEST,
            )
 
        if not batch_name:
            return Response(
                {"error": "Batch name is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
 
        filename = file.name.lower()
 
        # Only parse a 5-row preview here — the full file is read once,
        # later, inside Celery (process_bulk_upload). Reading the whole
        # thing twice was a chunk of the original slowness.
        try:
            if filename.endswith(".csv"):
                preview_df = pd.read_csv(file, dtype=str, keep_default_na=False, nrows=5)
            elif filename.endswith((".xlsx", ".xls")):
                preview_df = pd.read_excel(file, engine="openpyxl", dtype=str, nrows=5)
            else:
                return Response(
                    {"error": "Only CSV and Excel files are supported."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            return Response(
                {"error": f"Failed to read file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
 
        if preview_df.empty:
            return Response(
                {"error": "Uploaded file is empty."},
                status=status.HTTP_400_BAD_REQUEST,
            )
 
        # Cheap row count for display purposes only. The authoritative
        # count gets set again once process_bulk_upload actually parses
        # the full file.
        file.seek(0)
        total_records = 0
        try:
            if filename.endswith(".csv"):
                total_records = sum(1 for _ in file) - 1
            else:
                import openpyxl
                file.seek(0)
                wb = openpyxl.load_workbook(file, read_only=True)
                ws = wb.active
                total_records = max((ws.max_row or 1) - 1, 0)
                wb.close()
        except Exception:
            total_records = 0
 
        file.seek(0)
 
        exhibitor = get_exhibitor(request)
 
        batch = UploadBatch.objects.create(
            exhibitor=exhibitor,
            batch_name=batch_name,
            uploaded_file=file,
            file_name=file.name,
            total_records=total_records,
            status="uploaded",
        )
 
        columns = list(preview_df.columns)
        preview_rows = preview_df.fillna("").to_dict(orient="records")
 
        return Response(
            {
                "batch_id": batch.id,
                "batch_name": batch.batch_name,
                "file_name": batch.file_name,
                "total_records": batch.total_records,
                "columns": columns,
                "system_fields": SYSTEM_FIELDS,
                "preview_rows": preview_rows,
            },
            status=status.HTTP_201_CREATED,
        )
 
 
# ─────────────────────────────────────────────
# STEP 2 — Save field mapping, hand off to Celery
# POST /exhibitor/bulk-upload/<batch_id>/map/
# ─────────────────────────────────────────────
class BulkUploadMapFieldsView(APIView):
    permission_classes = [IsAuthenticated]
 
    def post(self, request, batch_id):
        batch = get_object_or_404(UploadBatch, id=batch_id)
 
        mappings = request.data.get("mappings", {})
 
        if not mappings:
            return Response(
                {"error": "Field mappings are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
 
        mapped_targets = set(mappings.values())
        missing = REQUIRED_TARGET_FIELDS - mapped_targets
 
        if missing:
            return Response(
                {"error": f"Missing required mappings: {', '.join(missing)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
 
        with transaction.atomic():
            UploadFieldMapping.objects.filter(batch=batch).delete()
 
            UploadFieldMapping.objects.bulk_create(
                [
                    UploadFieldMapping(
                        batch=batch, source_column=source, target_field=target
                    )
                    for source, target in mappings.items()
                ]
            )
 
            batch.status = "processing"
            batch.progress_percentage = 0
            batch.processed_records = 0
            batch.save(
                update_fields=["status", "progress_percentage", "processed_records"]
            )
 
        # Heavy lifting happens here, off the request/response cycle.
        process_bulk_upload.delay(batch.id, mappings)
 
        return Response(
            {
                "message": "Field mapping saved. Processing started.",
                "batch_id": batch.id,
                "status": batch.status,
            },
            status=status.HTTP_202_ACCEPTED,
        )
 
 
# ─────────────────────────────────────────────
# STEP 3 — Get batch status + records (with tabs)
# GET /exhibitor/bulk-upload/<batch_id>/review/
# ─────────────────────────────────────────────
class BulkUploadReviewView(APIView):
    permission_classes = [IsAuthenticated]
 
    def get(self, request, batch_id):
        batch = get_object_or_404(UploadBatch, id=batch_id)
 
        if batch.status not in ["validated", "completed"]:
            return Response(
                {
                    "batch_id": batch.id,
                    "batch_name": batch.batch_name,
                    "status": batch.status,
                    "progress_percentage": batch.progress_percentage,
                    "total_records": batch.total_records,
                    "processed_records": batch.processed_records,
                    "valid_records": batch.valid_records,
                    "invalid_records": batch.invalid_records,
                    "records": [],
                }
            )
 
        tab = request.query_params.get("tab", "all")
        search = request.query_params.get("search", "").strip()
 
        queryset = (
            batch.records.select_related("ticket_type").order_by("row_number")
        )
 
        if tab == "valid":
            queryset = queryset.filter(validation_status="valid")
        elif tab == "invalid":
            queryset = queryset.filter(validation_status="invalid")
 
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(email__icontains=search)
            )
 
        paginator = PageNumberPagination()
        paginator.page_size = min(int(request.query_params.get("page_size", 50)), 100)
 
        page = paginator.paginate_queryset(queryset, request)
        serializer = UploadBatchRecordSerializer(page, many=True)
 
        return paginator.get_paginated_response(
            {
                "batch_id": batch.id,
                "batch_name": batch.batch_name,
                "status": batch.status,
                "progress_percentage": batch.progress_percentage,
                "total_records": batch.total_records,
                "processed_records": batch.processed_records,
                "valid_records": batch.valid_records,
                "invalid_records": batch.invalid_records,
                "tab": tab,
                "records": serializer.data,
            }
        )
 
 
# ─────────────────────────────────────────────
# STEP 4 — Edit a single invalid record
# PATCH /exhibitor/bulk-upload/record/<record_id>/edit/
# ─────────────────────────────────────────────
class BulkUploadRecordEditView(APIView):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def patch(self, request, record_id):
        get_object_or_404(UploadBatchRecord.objects.select_for_update(), id=record_id)
        record = get_object_or_404(
            UploadBatchRecord.objects.select_related("batch__exhibitor", "ticket_type"),
            id=record_id,
        )

        old_status = record.validation_status

        serializer = RecordEditSerializer(record, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        record.refresh_from_db(fields=[
            "first_name", "last_name", "email",
            "ticket_type_id", "validation_status", "error_message",
        ])

        errors = []

        if not record.first_name:
            errors.append("First name is required.")
        if not record.last_name:
            errors.append("Last name is required.")
        if not record.email:
            errors.append("Email is required.")
        if not record.ticket_type_id:
            errors.append("Ticket type is required.")

        duplicate_in_batch = (
            UploadBatchRecord.objects.filter(
                batch=record.batch,
                email=record.email,
                validation_status="valid",
            )
            .exclude(id=record.id)
            .exists()
        )

        duplicate_committed = Registration.objects.filter(
            exhibitor=record.batch.exhibitor,
            email=record.email,
        ).exists()

        if duplicate_in_batch or duplicate_committed:
            errors.append("Duplicate email found.")

        # Check 4: global pool — unchanged
        batch_usage = 0
        if not errors and record.ticket_type_id:
            ticket_type_obj = record.ticket_type
            total_allowed = ticket_type_obj.total_tickets

            confirmed_count = Registration.objects.filter(
                ticket_type_id=record.ticket_type_id,
                status="confirmed",
            ).count()

            batch_usage = UploadBatchRecord.objects.filter(
                batch=record.batch,
                ticket_type_id=record.ticket_type_id,
                validation_status="valid",
            ).exclude(id=record.id).count()

            total_used = confirmed_count + batch_usage

            if total_used >= total_allowed:
                errors.append(
                    f"Badge quota exceeded for '{ticket_type_obj.ticket_name}' "
                    f"({total_used} used of {total_allowed} available). "
                    "Please contact your admin to allocate more badges "
                    "or choose a different ticket type."
                )

            # ✅ NEW — Check 5: this exhibitor's own allocation
            if not errors:
                exhibitor = record.batch.exhibitor
                allocated, exhibitor_confirmed, _ = get_exhibitor_allocation(
                    exhibitor, ticket_type_obj
                )
                exhibitor_used = exhibitor_confirmed + batch_usage
                exhibitor_available = allocated - exhibitor_used

                if allocated == 0 or exhibitor_available <= 0:
                    errors.append(
                        f"You have only {max(exhibitor_available, 0)} '{ticket_type_obj.ticket_name}' "
                        f"badge(s) remaining out of your allocated {allocated}. "
                        "Please contact the administrator to increase your quota."
                    )

        is_valid = len(errors) == 0
        new_status = "valid" if is_valid else "invalid"

        record.validation_status = new_status
        record.error_message = None if is_valid else " | ".join(errors)
        record.save(update_fields=["validation_status", "error_message"])

        if old_status != new_status:
            if old_status == "invalid" and new_status == "valid":
                UploadBatch.objects.filter(id=record.batch_id).update(
                    valid_records=F("valid_records") + 1,
                    invalid_records=F("invalid_records") - 1,
                )
            elif old_status == "valid" and new_status == "invalid":
                UploadBatch.objects.filter(id=record.batch_id).update(
                    valid_records=F("valid_records") - 1,
                    invalid_records=F("invalid_records") + 1,
                )

        record.batch.refresh_from_db(fields=["valid_records", "invalid_records"])

        return Response(
            {
                "message": "Record updated successfully",
                "record": UploadBatchRecordSerializer(record).data,
                "batch_valid_records": record.batch.valid_records,
                "batch_invalid_records": record.batch.invalid_records,
            }
        ) 
 
# ─────────────────────────────────────────────
# STEP 5 — Commit valid records to Registration
# POST /exhibitor/bulk-upload/<batch_id>/commit/
# ─────────────────────────────────────────────
class BulkUploadCommitView(APIView):
    permission_classes = [IsAuthenticated]
 
    @transaction.atomic
    def post(self, request, batch_id):
        batch = get_object_or_404(
            UploadBatch.objects.select_for_update(), id=batch_id
        )
 
        exhibitor = get_exhibitor(request)
 
        if batch.status == "completed":
            return Response(
                {"error": "Batch already committed."},
                status=status.HTTP_400_BAD_REQUEST,
            )
 
        if batch.status not in ("validated",):
            return Response(
                {
                    "error": (
                        "Batch is not ready for commit. "
                        f"Current status: {batch.status}."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
 
        if batch.valid_records == 0:
            return Response(
                {"error": "No valid records to commit."},
                status=status.HTTP_400_BAD_REQUEST,
            )
 
        # 1. Lock the status FIRST, inside the transaction.
        batch.status = "committing"
        batch.save(update_fields=["status"])
 
        # 2. Dispatch Celery ONLY after the transaction commits so the worker
        #    sees the status change before it starts.
        transaction.on_commit(
            lambda: commit_bulk_upload.delay(batch.id, exhibitor.id)
        )
 
        return Response(
            {
                "message": "Commit started.",
                "batch_id": batch.id,
                "valid_records": batch.valid_records,
            }
        )
 
 
# ─────────────────────────────────────────────
# List all batches for exhibitor
# GET /exhibitor/bulk-upload/batches/
# ─────────────────────────────────────────────
class BulkUploadBatchListView(APIView):
    permission_classes = [IsAuthenticated]
 
    def get(self, request):
        exhibitor = get_exhibitor(request)
        batches = UploadBatch.objects.filter(exhibitor=exhibitor).order_by("-uploaded_at")
        serializer = UploadBatchListSerializer(batches, many=True)
        return Response(serializer.data)
 
 
# ─────────────────────────────────────────────
# Sample template download
# GET /exhibitor/bulk-upload/sample-template/
# ─────────────────────────────────────────────
class BulkUploadSampleTemplateView(APIView):
    permission_classes = [IsAuthenticated]
 
    def get(self, request):
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Alignment, Font, PatternFill
 
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registrations"
 
        headers = [
            "First Name", "Last Name", "Email", "Job Title", "Company Name",
            "Phone Number", "Country Of Residence", "Nationality", "Ticket Type",
        ]
 
        header_font = Font(bold=True, color="FFFFFF", name="Arial")
        header_fill = PatternFill("solid", start_color="3f0e60")
 
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
 
        ws.append([
            "John", "Doe", "john.doe@example.com", "Manager", "ABC Company",
            "+971501234567", "United Arab Emirates", "Emirati", "Exhibitor Badge",
        ])
 
        col_widths = [15, 15, 30, 20, 25, 18, 25, 20, 20]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
 
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
 
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="bulk_upload_template.xlsx"'
        return response
 
 
# ─────────────────────────────────────────────
# Bulk delete registrations by ID list
# DELETE /exhibitor/registrations/bulk-delete/
# Body: { "ids": [1, 2, 3] }
# ─────────────────────────────────────────────
class RegistrationBulkDeleteView(APIView):
    permission_classes = [IsAuthenticated]
 
    def delete(self, request):
        ids = request.data.get("ids", [])
 
        if not ids:
            return Response({"error": "No IDs provided."}, status=status.HTTP_400_BAD_REQUEST)
        if not isinstance(ids, list):
            return Response({"error": "ids must be a list."}, status=status.HTTP_400_BAD_REQUEST)
        if len(ids) > 500:
            return Response({"error": "Cannot delete more than 500 records at once."}, status=status.HTTP_400_BAD_REQUEST)
 
        exhibitor = get_exhibitor(request)
 
        qs = Registration.objects.filter(id__in=ids, exhibitor=exhibitor)
 
        # Capture affected batches BEFORE deleting rows
        affected_batch_ids = list(
            qs.exclude(upload_batch__isnull=True)
              .values_list("upload_batch_id", flat=True)
              .distinct()
        )
 
        requested_count = len(ids)
        deleted_count, _ = qs.delete()
 
        # Reconcile: if a "completed" batch has no registrations left, flip it back
        if affected_batch_ids:
            for batch in UploadBatch.objects.filter(id__in=affected_batch_ids, status="completed"):
                if not Registration.objects.filter(upload_batch=batch, exhibitor=exhibitor).exists():
                    batch.status = "validated"
                    batch.save(update_fields=["status"])
 
        return Response(
            {
                "message": f"{deleted_count} registration(s) deleted.",
                "deleted_count": deleted_count,
                "requested_count": requested_count,
            },
            status=status.HTTP_200_OK,
        )
 
 
# ─────────────────────────────────────────────
# Delete all registrations for a committed batch
# DELETE /exhibitor/bulk-upload/<batch_id>/delete-registrations/
# ─────────────────────────────────────────────
class BulkUploadDeleteRegistrationsView(APIView):
    permission_classes = [IsAuthenticated]
 
    def delete(self, request, batch_id):
        exhibitor = get_exhibitor(request)
        batch = get_object_or_404(UploadBatch, id=batch_id, exhibitor=exhibitor)
 
        if batch.status != "completed":
            return Response(
                {"error": "This batch has not been committed yet."},
                status=status.HTTP_400_BAD_REQUEST,
            )
 
        deleted_count, _ = Registration.objects.filter(
            upload_batch=batch, exhibitor=exhibitor,
        ).delete()
 
        if deleted_count == 0:
            return Response(
                {"error": "No registrations found for this batch. They may have already been deleted."},
                status=status.HTTP_400_BAD_REQUEST,
            )
 
        batch.status = "validated"
        batch.save(update_fields=["status"])
 
        return Response({
            "message": f"{deleted_count} registrations deleted.",
            "deleted_count": deleted_count,
        })
 



class RegistrationCreateAPIView(CreateAPIView):
    serializer_class = RegistrationCreateSerializer
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def perform_create(self, serializer):
        exhibitor = self.request.user.exhibitor_profile
        ticket_type = serializer.validated_data["ticket_type"]

        allocated, exhibitor_used, exhibitor_available = get_exhibitor_allocation(
            exhibitor, ticket_type
        )

        if allocated == 0:
            raise ValidationError(
                {
                    "ticket_type": (
                        f"You have not been allocated any '{ticket_type.ticket_name}' "
                        "badges. Please contact the administrator."
                    )
                }
            )

        if exhibitor_available <= 0:
            raise ValidationError(
                {
                    "ticket_type": (
                        f"You have only {exhibitor_available} '{ticket_type.ticket_name}' "
                        f"badge(s) remaining out of your allocated {allocated}. "
                        "Please contact the administrator to increase your quota."
                    )
                }
            )

        used_count = Registration.objects.filter(
            ticket_type=ticket_type
        ).exclude(status="cancelled").count()

        if ticket_type.total_tickets - used_count <= 0:
            raise ValidationError(
                {"ticket_type": f"No available badges left for '{ticket_type.ticket_name}'."}
            )

        serializer.save(
            exhibitor=exhibitor,
            status="confirmed",
            registered_via="single_badge",
        )
        
class RegistrationUpdateAPIView(UpdateAPIView):

    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    serializer_class = RegistrationUpdateSerializer

    def get_queryset(self):
        exhibitor = self.request.user.exhibitor_profile
        return Registration.objects.filter(exhibitor=exhibitor)

    def perform_update(self, serializer):
        registration = serializer.instance  # already-fetched instance, no extra query
        new_ticket_type = serializer.validated_data.get(
            "ticket_type", registration.ticket_type
        )

        # ✅ NEW — only re-check quota when the ticket type is actually changing.
        # If unchanged, this registration already holds its seat.
        if new_ticket_type != registration.ticket_type:
            exhibitor = registration.exhibitor

            allocated, exhibitor_used, exhibitor_available = get_exhibitor_allocation(
                exhibitor, new_ticket_type
            )

            if allocated == 0:
                raise ValidationError(
                    {
                        "ticket_type": (
                            f"You have not been allocated any '{new_ticket_type.ticket_name}' "
                            "badges. Please contact the administrator."
                        )
                    }
                )

            if exhibitor_available <= 0:
                raise ValidationError(
                    {
                        "ticket_type": (
                            f"You have only {exhibitor_available} '{new_ticket_type.ticket_name}' "
                            f"badge(s) remaining out of your allocated {allocated}. "
                            "Please contact the administrator to increase your quota."
                        )
                    }
                )

            # Global pool check — exclude this registration's own row since
            # it's about to move INTO this type, not currently counted against it.
            used_count = Registration.objects.filter(
                ticket_type=new_ticket_type
            ).exclude(
                status="cancelled"
            ).exclude(
                id=registration.id
            ).count()

            available_count = new_ticket_type.total_tickets - used_count

            if available_count <= 0:
                raise ValidationError(
                    {
                        "ticket_type": (
                            f"No available badges left for '{new_ticket_type.ticket_name}'."
                        )
                    }
                )

        serializer.save()


class RegistrationDeleteAPIView(DestroyAPIView):

    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):

        exhibitor = self.request.user.exhibitor_profile

        return Registration.objects.filter(
            exhibitor=exhibitor
        )





class RegistrationListAPIView(ListAPIView):

    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    serializer_class = RegistrationListSerializer
    pagination_class = None   # ← disable pagination, return everything in one response

    def get_queryset(self):

        exhibitor = (self.request.user.exhibitor_profile)

        queryset = (
            Registration.objects.filter(exhibitor=exhibitor)

            .select_related("ticket_type", "upload_batch", "invitation")

            .order_by("-created_at")
        )

        ticket = self.request.GET.get("ticket")
        status = self.request.GET.get("status")

        if ticket:
            queryset = queryset.filter(ticket_type_id=ticket)

        if status:
            queryset = queryset.filter(status=status)

        return queryset
        
class ExhibitorDashboardAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        exhibitor = request.user.exhibitor_profile

        # ── Status counts ─────────────────────────────────────────────────────
        status_counts = (
            Registration.objects.filter(exhibitor=exhibitor)
            .values("status")
            .annotate(count=Count("id"))
        )
        counts    = {item["status"]: item["count"] for item in status_counts}
        confirmed = counts.get("confirmed", 0)
        pending   = counts.get("pending",   0)
        invited   = counts.get("invited",   0)

        # ── Per-ticket-type breakdown ─────────────────────────────────────────
        # Prefer BadgeAllocation if it exists, fall back to TicketType.total_tickets
        allocations = (
            BadgeAllocation.objects.filter(exhibitor=exhibitor)
            .select_related("ticket_type")
        )

        ticket_breakdown = []
        total_allocated  = 0

        if allocations.exists():
            # Admin has explicitly allocated badges — use those numbers
            for alloc in allocations:
                tt   = alloc.ticket_type
                used = Registration.objects.filter(
                    exhibitor=exhibitor,
                    ticket_type=tt,
                ).exclude(status="cancelled").count()

                allocated     = alloc.allocated_count
                total_allocated += allocated

                ticket_breakdown.append({
                    "id":          tt.id,
                    "ticket_name": tt.ticket_name,
                    "allocated":   allocated,
                    "used":        used,
                    "available":   max(allocated - used, 0),
                })

        else:
            # No BadgeAllocation yet — fall back to TicketType.total_tickets
            ticket_types = TicketType.objects.filter(status="active")

            for tt in ticket_types:
                used = Registration.objects.filter(
                    exhibitor=exhibitor,
                    ticket_type=tt,
                ).exclude(status="cancelled").count()

                allocated     = tt.total_tickets
                total_allocated += allocated

                ticket_breakdown.append({
                    "id":          tt.id,
                    "ticket_name": tt.ticket_name,
                    "allocated":   allocated,
                    "used":        used,
                    "available":   max(allocated - used, 0),
                })

        return Response({
            "allocated_badges": total_allocated,
            "confirmed":        confirmed,
            "pending":          pending,
            "invited":          invited,
            "available_badges": max(total_allocated - confirmed, 0),
            "ticket_breakdown": ticket_breakdown,
        })
                
class ExhibitorLoginAPIView(APIView):

    permission_classes = []

    def post(self, request):

        # ✅ Verify captcha before touching credentials at all
        captcha_token = request.data.get("captcha_token")
        if not verify_recaptcha(captcha_token):
            return Response(
                {
                    "status": False,
                    "message": "Captcha verification failed. Please try again."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = (ExhibitorLoginSerializer(data=request.data))

        serializer.is_valid(raise_exception=True)

        username = (serializer.validated_data["username"])

        password = (serializer.validated_data["password"])

        user = authenticate(
            username=username,
            password=password
        )

        if not user:

            return Response(
                {
                    "status": False,
                    "message":
                    "Invalid credentials"
                },
                status=status.HTTP_401_UNAUTHORIZED
            )

        if user.role != "exhibitor":

            return Response(
                {
                    "status": False,
                    "message":
                    "Not an exhibitor account"
                },
                status=status.HTTP_403_FORBIDDEN
            )

        refresh = (RefreshToken.for_user(user))

        return Response(
            {
                "status": True,

                "access":str(refresh.access_token),

                "refresh":str(refresh),

                "user": {"id":user.id,

                    "username":user.username,

                    "role":user.role
                }
            }
        )

class ExhibitorDetailAPIView(RetrieveAPIView):

    queryset = (Exhibitor.objects.all())

    serializer_class = (ExhibitorDetailSerializer)

class ExhibitorListAPIView(ListAPIView):

    serializer_class = (ExhibitorListSerializer)

    queryset = (Exhibitor.objects.all().order_by("-id"))


class CreateExhibitorAPIView(APIView):

    @transaction.atomic
    def post(self, request):

        serializer = CreateExhibitorSerializer(data=request.data)

        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        user = User.objects.create_user(
            username=data["username"],
            email=data["contact_email"],
            password=data["password"]
        )

        user.role = "exhibitor"
        user.save()

        exhibitor = Exhibitor.objects.create(
            user=user,
            company_name=data["company_name"],
            contact_person=data["contact_person"],
            contact_email=data["contact_email"],
            contact_phone=data["contact_phone"]
        )

        return Response(
            {
                "status": True,
                "message":
                "Exhibitor created successfully",
                "id": exhibitor.id
            }
        )

class TicketTypeDetailAPIView(RetrieveAPIView):

    queryset = (TicketType.objects.all())

    serializer_class = (TicketTypeListSerializer)

class DeleteTicketTypeAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        ticket = get_object_or_404(TicketType, id=pk)

        # ✅ Block deactivation if any active registrations exist
        active_count = Registration.objects.filter(
            ticket_type=ticket
        ).exclude(
            status="cancelled"
        ).count()

        if active_count > 0:
            return Response(
                {
                    "error": (
                        f"Cannot deactivate '{ticket.ticket_name}'. "
                        f"{active_count} badge(s) are already registered against this ticket."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        ticket.status = "inactive"
        ticket.save()

        return Response(
            {"message": "Ticket deactivated"},
            status=status.HTTP_200_OK,
        )


class UpdateTicketTypeAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        ticket = get_object_or_404(TicketType, id=pk)
        serializer = TicketTypeUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        # ✅ Block reducing total_tickets below already-used count
        if "total_tickets" in data:
            used_count = Registration.objects.filter(
                ticket_type=ticket
            ).exclude(
                status="cancelled"
            ).count()

            if data["total_tickets"] < used_count:
                return Response(
                    {
                        "error": (
                            f"Cannot reduce total tickets to {data['total_tickets']}. "
                            f"{used_count} badges are already registered against this ticket."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        for field, value in data.items():
            setattr(ticket, field, value)
        ticket.save()

        return Response(
            {
                "message": "Ticket updated",
                "ticket": {
                    "id": ticket.id,
                    "ticket_name": ticket.ticket_name,
                    "ticket_code": ticket.ticket_code,
                    "total_tickets": ticket.total_tickets,
                    "description": ticket.description,
                    "status": ticket.status,
                }
            },
            status=status.HTTP_200_OK,
        )


class DeleteExhibitorAPIView(APIView):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def delete(self, request, pk):
        exhibitor = get_object_or_404(Exhibitor, id=pk)

        registration_count = Registration.objects.filter(exhibitor=exhibitor).count()
        badge_count = Registration.objects.filter(exhibitor=exhibitor).exclude(status="cancelled").count()
        invitation_count = Invitation.objects.filter(registration__exhibitor=exhibitor).count()

        user = exhibitor.user
        exhibitor.delete()
        user.delete()

        return Response({
            "status": True,
            "message": "Exhibitor deleted successfully.",
            "removed": {
                "registrations": registration_count,
                "active_badges": badge_count,
                "invitations": invitation_count,
            }
        }, status=status.HTTP_200_OK)


class TicketTypeListAPIView(ListAPIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    serializer_class = TicketTypeListSerializer

    def get_queryset(self):
        return TicketType.objects.all()



class CreateTicketTypeAPIView(APIView):

    def post(self, request):

        serializer = TicketTypeCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        ticket = TicketType.objects.create(
            ticket_name=data["ticket_name"],
            ticket_code=data["ticket_code"],
            total_tickets=data["total_tickets"],
            description=data.get("description", ""),
            status="active",   # ← explicit, don't rely on a model/db default
        )

        return Response(
            {"message": "Ticket type created", "id": ticket.id},
            status=201
        )

class AllocateBadgeAPIView(APIView):

    def post(self, request):

        serializer = (BadgeAllocationSerializer(data=request.data))

        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        exhibitor = Exhibitor.objects.get(id=data["exhibitor_id"])

        ticket = TicketType.objects.get(
            id=data["ticket_type_id"]
        )

        allocation, created = (
            BadgeAllocation.objects.update_or_create(
                exhibitor=exhibitor,
                ticket_type=ticket,
                defaults={
                    "allocated_count":data["allocated_count"],

                    "allocated_by":request.user,

                    "remarks":data.get("remarks")
                }
            )
        )

        return Response({
            "message":
            "Allocation successful"
        })


class SuperAdminLogoutView(APIView):
    def post(self, request):
        try:
            refresh_token = request.data.get("refresh_token")
            token = RefreshToken(refresh_token)
            token.blacklist()
            return Response(
                {
                    "status": True,
                    "message": "Logout Successful"
                },
                status=status.HTTP_205_RESET_CONTENT
            )
        except Exception as e:
            return Response(
                {
                    "status": False,
                    "message": "Invalid Token"
                },
                status=status.HTTP_400_BAD_REQUEST
            )

class SuperAdminLoginView(APIView):

    authentication_classes = []
    permission_classes = []

    def post(self, request):

        # ✅ Verify captcha before touching credentials at all
        captcha_token = request.data.get("captcha_token")
        if not verify_recaptcha(captcha_token):
            return Response(
                {
                    "status": False,
                    "message": "Captcha verification failed. Please try again."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = LoginSerializer(data=request.data)

        if not serializer.is_valid():
            return Response(
                serializer.errors,
                status=status.HTTP_400_BAD_REQUEST
            )

        username = serializer.validated_data['username']
        password = serializer.validated_data['password']

        user = authenticate(
            username=username,
            password=password
        )

        if not user:
            return Response(
                {
                    "status": False,
                    "message": "Invalid Username or Password"
                },
                status=status.HTTP_401_UNAUTHORIZED
            )

        if not user.is_superuser and getattr(user, "role", None) == "exhibitor":
            return Response(
                {
                    "status": False,
                    "message": "You don't have the permission to access this side"
                },
                status=status.HTTP_401_UNAUTHORIZED
            )

        refresh = RefreshToken.for_user(user)

        return Response({
            "status": True,
            "message": "Login Successful",
            "access_token": str(refresh.access_token),
            "refresh_token": str(refresh),
            "user": {
                "id": user.id,
                "username": user.username,
                "email": user.email,
                "role": getattr(user, "role", None),
                "is_superuser": user.is_superuser
            }
        })
